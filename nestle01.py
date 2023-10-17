import sqlite3
import openpyxl


def start():
    con = sqlite3.connect('nestle.db')

def post_staff():
    try:
        con = sqlite3.connect('nestle.db')
        con.execute("""CREATE TABLE IF NOT EXISTS poststaff(
        id INTEGER PRIMARY KEY AUTOINCREMENT,    
        title TEXT);
        """)
        
    except FileNotFoundError as e:
        print(f'Ошибка --> {e}')
    else:
        print('Создание таблицы poststaff завершенно')
    finally:
        con.commit()
        print('соеденение закрыто')

def staff():
    try:
        con = sqlite3.connect('nestle.db')
        con.execute("""CREATE TABLE IF NOT EXISTS staff(
        id INTEGER PRIMARY KEY AUTOINCREMENT,    
        name TEXT,
        post TEXT,
        FOREIGN KEY (post)  REFERENCES poststaff (id));
        """)
        
    except FileNotFoundError as e:
        print(f'Ошибка --> {e}')
    else:
        print('Создание таблицы staff завершенно')
    finally:
        con.commit()
        print('соеденение закрыть')

def customers():
    try:
        con = sqlite3.connect('nestle.db')
        con.execute("""CREATE TABLE IF NOT EXISTS customers(
        id INTEGER PRIMARY KEY AUTOINCREMENT,    
        title TEXT,
        inn TEXT,
        address TEXT);
        """)
        
    except FileNotFoundError as e:
        print(f'Ошибка --> {e}')
    else:
        print('Создание таблицы customers завершенно')
    finally:
        con.commit()
        print('соеденение закрыто')

def products():
    try:
        con = sqlite3.connect('nestle.db')
        con.execute("""CREATE TABLE IF NOT EXISTS products(
        id INTEGER PRIMARY KEY AUTOINCREMENT,    
        title TEXT,
        ediz TEXT,
        article TEXT,
        price REAL);
        """)
        
    except FileNotFoundError as e:
        print(f'Ошибка --> {e}')
    else:
        print('Создание таблицы products завершенно')
    finally:
        con.commit()
        print('соеденение закрыто')

def invoice():
    try:
        con = sqlite3.connect('nestle.db')
        con.execute("""CREATE TABLE IF NOT EXISTS invoice(
        id INTEGER PRIMARY KEY AUTOINCREMENT,    
        score TEXT,                
        products TEXT,        
        kolvo INTEGER,        
        FOREIGN KEY (score)  REFERENCES score (id),
        FOREIGN KEY (products)  REFERENCES products (id));
        """)
        
    except FileNotFoundError as e:
        print(f'Ошибка --> {e}')
    else:
        print('Создание таблицы invoice завершенно')
    finally:
        con.commit()
        print('соеденение закрыто')
def score():
    try:
        con = sqlite3.connect('nestle.db')
        con.execute("""CREATE TABLE IF NOT EXISTS score(
        id INTEGER PRIMARY KEY AUTOINCREMENT,    
        title TEXT,
        staff TEXT,        
        customers TEXT,      
        data TEXT,
        FOREIGN KEY (staff)  REFERENCES staff (id),
        FOREIGN KEY (customers)  REFERENCES customers (id));
        """)
        
    except FileNotFoundError as e:
        print(f'Ошибка --> {e}')
    else:
        print('Создание таблицы score завершенно')
    finally:
        con.commit()
        print('соеденение закрыто')


def customers_add(fail: str):

    
    wookbook = openpyxl.load_workbook(f"{fail}")

    
    worksheet = wookbook.active

    x = 0
    for i in range(0, worksheet.max_row):
        se = []
        x += 1
        for col in worksheet.iter_cols(1, worksheet.max_column):
            #print(col[i].value, end="\t\t")
            se += [col[i].value]
            #print(se, 'это нужный кортеж')
        print('')
        print(tuple(se))
        se1 = tuple(se)
        if x > 1:
            try:
                con = sqlite3.connect('nestle.db')
                con.execute('INSERT INTO customers(title, inn, address) VALUES(?, ?, ?)', se1)
            except FileNotFoundError as e:
                print(f'Ошибка --> {e}')
            else:
                #print('Значение добавленно', se1)
                pass
            finally:
                con.commit()
                #print('соеденение закрыто')

def staff_add(fail: str):

    
    wookbook = openpyxl.load_workbook(f"{fail}")

    
    worksheet = wookbook.active

    x = 0
    for i in range(0, worksheet.max_row):
        se = []
        x += 1
        for col in worksheet.iter_cols(1, worksheet.max_column):
            #print(col[i].value, end="\t\t")
            se += [col[i].value]
            print(se, 'это нужный кортеж')
        print('')
        print(tuple(se))
        se1 = tuple(se)
        if x > 1:
            try:
                con = sqlite3.connect('nestle.db')
                con.execute('INSERT INTO staff(name, post) VALUES(?, ?)', se1)
            except FileNotFoundError as e:
                print(f'Ошибка --> {e}')
            else:
                print('Значение добавленно', se1)
                
            finally:
                con.commit()
                #print('соеденение закрыто')

def products_add(fail: str):

    
    wookbook = openpyxl.load_workbook(f"{fail}")

    
    worksheet = wookbook.active

    x = 0
    for i in range(0, worksheet.max_row):
        se = []
        x += 1
        for col in worksheet.iter_cols(1, worksheet.max_column):
            #print(col[i].value, end="\t\t")
            se += [col[i].value]
            #print(se, 'это нужный кортеж')
        print('')
        print(tuple(se))
        se1 = tuple(se)
        if x > 1:
            try:
                con = sqlite3.connect('nestle.db')
                con.execute('INSERT INTO products(title, ediz, article, price) VALUES(?, ?, ?, ?)', se1)
            except FileNotFoundError as e:
                print(f'Ошибка --> {e}')
            else:
                #print('Значение добавленно', se1)
                pass
            finally:
                con.commit()
                #print('соеденение закрыто')


def post_staff_add(fail: str):
    try:
        con = sqlite3.connect('nestle.db')
        con.execute('INSERT INTO poststaff(title) VALUES(?)', (fail,))
    except FileNotFoundError as e:
        print(f'Ошибка --> {e}')
    else:
        print('Значение добавленно', fail)
        cursor = con.cursor()
        sql = "SELECT * FROM poststaff WHERE title=?"
        cursor.execute(sql, (fail,))
        #rec = con.execute(f"SELECT title from poststaff where title = {fail}")
        
        se = cursor.fetchall()
        print('id сущности', se[0][0])
    finally:
        con.commit()
        print('соеденение закрыто')

class Score_add():
    def __init__(self,title, staff, customers,data):
        self.title = title
        self.staff = staff
        self.customers = customers        
        self.data = data
        self.id = 0
        def ad1(title, staff, customers, data):
            try:
                se1 = (title, staff, customers, data)
                con = sqlite3.connect('nestle.db')
                cursor = con.cursor()
                cursor.execute('INSERT INTO score(title, staff, customers, data) VALUES(?, ?, ?, ?)', se1)
            except FileNotFoundError as e:
                print(f'Ошибка --> {e}')
            else:
                sql = "SELECT * FROM score WHERE title=?"
                cursor.execute(sql, (title,))
                #rec = con.execute(f"SELECT title from customers where id = {fail}")
                se5 = cursor.fetchall()
                if len(se5) == 1:
                                    
                    print('Счет создан', se1[0], 'id счета ', se5[0][0])
                    self.id = se5[0][0]
                    
                else:
                    print("длина", len(se5))
                    x2 = 0
                    for i in se5:
                        if x2 > 0:
                            h = int(i[0])
                            print(h, type(h))
                            
                            h1 = (h)
                            mydata = cursor.execute("DELETE FROM score WHERE id=?", (h1,))
                        x2 += 1    

                    print('Не уникальное наименнованиет счета')
                

                
            finally:
                con.commit()
                #print('соеденение закрыто')
        ad1(title, staff, customers, data)
    def add_in(self, products: int, kolvo: int):
        if self.id == 0:
            return 'Счет не уникальный позицию в счете не добавить'
        try:
            con = sqlite3.connect('nestle.db')
            cursor = con.cursor()
            sql = "SELECT * FROM products WHERE id=?"
            cursor.execute(sql, (products,))
            #rec = con.execute(f"SELECT title from customers where id = {fail}")
            se = cursor.fetchall()
            if len(se) == 0:
                return 'нет товара с данным id'
            
            con.commit()
            
        except:
            con.commit()
            
        try:
            score = int(self.id)
            se1 = (score, products, kolvo)
            con = sqlite3.connect('nestle.db')
            cursor = con.cursor()
            cursor.execute('INSERT INTO invoice(score, products, kolvo) VALUES(?, ?, ?)', se1)
        except FileNotFoundError as e:
                print(f'Ошибка --> {e}')
        else:
            print(f'Позиция добавленна в счет-->  {products} {kolvo}')
        finally:
            con.commit()


    def __str__(self):                  
               
        return f'счет {self.title}'


class Score_true():
    def __init__(self, id: int):
        self.id = id
    def __str__(self):
        
        try:
            
            con = sqlite3.connect('nestle.db')
            cursor = con.cursor()
            sql = "SELECT * FROM score WHERE id=?"
            cursor.execute(sql, (self.id,))          
            se5 = cursor.fetchall()
            se6 = se5[0][1]          
            
            
            con.commit()
            return f'Счет {se6}'
        except:
            con.commit()
            return f'Нет счета с id {self.id}'

    def add(self, products: int, kolvo: int):




        try:
            
            con = sqlite3.connect('nestle.db')
            cursor = con.cursor()
            sql = "SELECT * FROM score WHERE id=?"
            cursor.execute(sql, (self.id,))          
            se5 = cursor.fetchall()
            se6 = se5[0][1]          
            
            
            con.commit()
            t = 1
        except:
            con.commit()
            t = 0
            




        if t == 1:
            try:
                con = sqlite3.connect('nestle.db')
                cursor = con.cursor()
                sql = "SELECT * FROM products WHERE id=?"
                cursor.execute(sql, (products,))
                #rec = con.execute(f"SELECT title from customers where id = {fail}")
                se = cursor.fetchall()
                if len(se) == 0:
                    return 'нет товара с данным id'
                
                con.commit()
                
            except:
                con.commit()
                
            try:
                score = int(self.id)
                se1 = (score, products, kolvo)
                con = sqlite3.connect('nestle.db')
                cursor = con.cursor()
                cursor.execute('INSERT INTO invoice(score, products, kolvo) VALUES(?, ?, ?)', se1)
            except FileNotFoundError as e:
                    print(f'Ошибка --> {e}')
            else:
                print(f'Позиция добавленна в счет-->  {products} {kolvo}')
            finally:
                con.commit()
    def get(self):
            r = self.id
            #print(self.id)
            con = sqlite3.connect('nestle.db')
            cursor = con.cursor()
            sql = "SELECT * FROM invoice WHERE score=?"
            cursor.execute(sql, (r,))
            #rec = con.execute(f"SELECT title from customers where id = {fail}")
            se5 = cursor.fetchall()
            #print(se5)
            #print(len(se5))
            if len(se5) > 0:
                wer = 0
                for i in se5:
                    #print(i)
                    id = int(i[2])
                    #print('товар', id)                    
                    con = sqlite3.connect('nestle.db')
                    cursor = con.cursor()
                    sql = "SELECT * FROM products WHERE id=?"
                    cursor.execute(sql, (id,))
                    #rec = con.execute(f"SELECT title from customers where id = {fail}")
                    se = cursor.fetchall()
                    #print('se', se)
                    if len(se) == 0:
                        print('нет товара с данным id')
                        
                        con.commit()
                    else:
                        s = se[0][1]
                        s1 = se[0][4]
                        #print(3)
                        s2 = s1 * i[3]
                        print(f'id {i[0]} | {s} стоимость за ед. {s1} | кол-во {i[3]} | стоимость {s2} руб.')
                        wer += s2
                        con.commit()
                print(f'Общая стоимсоть {wer} руб.')
    def in_del(self, id: int):
        h1 = id
        con = sqlite3.connect('nestle.db')
        cursor = con.cursor()
        mydata = cursor.execute("DELETE FROM invoice WHERE id=?", (h1,))
        con.commit()
        print(f'Позиция {h1} удалена')


class Search():
    def search_staff(self, id: int):
        
        con = sqlite3.connect('nestle.db')
        cursor = con.cursor()
        sql = "SELECT * FROM score WHERE staff=?"
        cursor.execute(sql, (id,))
        
        se = cursor.fetchall()
        #se0 =se[0][1]
        #print(se, len(se))
        wer = 0
        for i in se:
            #print(i)
            #print(i[0], i[1])
            t = i[1]

            con = sqlite3.connect('nestle.db')
            cursor = con.cursor()
            sql = "SELECT * FROM invoice WHERE score=?"
            cursor.execute(sql, (int(i[0]),))
            #rec = con.execute(f"SELECT title from customers where id = {fail}")
            se5 = cursor.fetchall()

            ##########
            con = sqlite3.connect('nestle.db')
            cursor = con.cursor()
            sql = "SELECT * FROM customers WHERE id=?"
            cursor.execute(sql, (int(i[3]),))
            #rec = con.execute(f"SELECT title from customers where id = {fail}")
            se7 = cursor.fetchall()
            t2 = se7[0][1]
            ##########


            ##########
            con = sqlite3.connect('nestle.db')
            cursor = con.cursor()
            sql = "SELECT * FROM staff WHERE id=?"
            cursor.execute(sql, (int(i[2]),))
            #rec = con.execute(f"SELECT title from customers where id = {fail}")
            se6 = cursor.fetchall()
            t1 = se6[0][1]
            ##########




            if len(se5) > 0:
                    
                    for i in se5:
                        #print(i)
                        id = int(i[2])
                        #print(id)
                        #print('товар', id)                    
                        con = sqlite3.connect('nestle.db')
                        cursor = con.cursor()
                        sql = "SELECT * FROM products WHERE id=?"
                        cursor.execute(sql, (id,))
                        #rec = con.execute(f"SELECT title from customers where id = {fail}")
                        se = cursor.fetchall()
                        #print('se', se)
                        if len(se) == 0:
                            print('нет товара с данным id')
                            
                            con.commit()
                        else:
                            s = se[0][1]
                            s1 = se[0][4]
                            #print(3)
                            s2 = s1 * i[3]
                            print(f' {t1}|  {t2} | {t} |  {s} |  {i[3]} |  {s2} руб.')
                            #print(f'{se0} id  | {s} стоимость за ед. {s1} | кол-во {i[3]} | стоимость {s2} руб.')
                            wer += s2
                            con.commit()
                    
                
                    
            con.commit()
        print(f'Общая стоимсоть {wer} руб.')
    def search_staff_01(self, id: int):
        
        con = sqlite3.connect('nestle.db')
        cursor = con.cursor()
        sql = "SELECT * FROM score WHERE staff=?"
        cursor.execute(sql, (id,))
        
        se = cursor.fetchall()
        #se0 =se[0][1]
        #print(se, len(se))
        wer = 0
        for i in se:
            #print(i)
            #print(i[0], i[1])
            t = i[1]

            con = sqlite3.connect('nestle.db')
            cursor = con.cursor()
            sql = "SELECT * FROM invoice WHERE score=?"
            cursor.execute(sql, (int(i[0]),))
            #rec = con.execute(f"SELECT title from customers where id = {fail}")
            se5 = cursor.fetchall()

            ##########
            con = sqlite3.connect('nestle.db')
            cursor = con.cursor()
            sql = "SELECT * FROM customers WHERE id=?"
            cursor.execute(sql, (int(i[3]),))
            #rec = con.execute(f"SELECT title from customers where id = {fail}")
            se7 = cursor.fetchall()
            t2 = se7[0][1]
            ##########


            ##########
            con = sqlite3.connect('nestle.db')
            cursor = con.cursor()
            sql = "SELECT * FROM staff WHERE id=?"
            cursor.execute(sql, (int(i[2]),))
            #rec = con.execute(f"SELECT title from customers where id = {fail}")
            se6 = cursor.fetchall()
            t1 = se6[0][1]
            ##########




            if len(se5) > 0:
                    
                    for i in se5:
                        #print(i)
                        id = int(i[2])
                        #print(id)
                        #print('товар', id)                    
                        con = sqlite3.connect('nestle.db')
                        cursor = con.cursor()
                        sql = "SELECT * FROM products WHERE id=?"
                        cursor.execute(sql, (id,))
                        #rec = con.execute(f"SELECT title from customers where id = {fail}")
                        se = cursor.fetchall()
                        #print('se', se)
                        if len(se) == 0:
                            print('нет товара с данным id')
                            
                            con.commit()
                        else:
                            s = se[0][1]
                            s1 = se[0][4]
                            #print(3)
                            s2 = s1 * i[3]
                            print(f' {t1}|  {t2} |  {s} |  {i[3]} |  {s2} руб.')
                            #print(f'{se0} id  | {s} стоимость за ед. {s1} | кол-во {i[3]} | стоимость {s2} руб.')
                            wer += s2
                            con.commit()
                    
                
                    
            con.commit()
        print(f'Общая стоимсоть {wer} руб.')



    
            
        

    
    
        
        



if __name__ == "__main__":
    ####Создаем базу и таблицы #####
    start()
    post_staff()
    staff()
    customers()
    products()
    score()
    invoice()
    ####Создание базу и таблиц завершенно #####

    #### Импорт данных ####
    customers_add("customers.xlsx")
    post_staff_add('Менеджер')
    staff_add("staff.xlsx")
    products_add('products.xlsx')
    ##### Завершели импорт данных ######

